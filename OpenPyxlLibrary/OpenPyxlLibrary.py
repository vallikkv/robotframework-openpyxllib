#!/usr/bin/env python



import openpyxl

class  OpenPyxlLibrary:
    """
        This test library internally use openpyxl module of python and provides keywords to open, read, write excel files. This library only supports
        xlsx file formats.


        *Prerequisties*

        Openpyxl module of python should be installed using command "pip install openpyxl"
        OpenPyxlLibrary must be imported.

        Example:
            | Library        | OpenpyxlLibrary        |
            | Open Excel     | Filename with fullpath |

        """

    def __init__(self):
        self.wb = None
        self.sheet = None
        self.filename = None
        
    def open_excel(self, file):
        """
        Open excel file
        Arguments:
            | File             | Filename with fullpath to open and test upon        |

        Example:
        | Open Excel      |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |
        """
        self.filename = file
        self.wb = openpyxl.load_workbook(self.filename)

    def get_sheet_names(self):
        """
        Return sheetnames of the workbook
        Example:
        | Openexcel File       |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |
        | Get sheet names      |                                                     |
        """
        self.filename = file
        return self.wb.get_sheet_names()
    
    
    def opensheet_byname(self, sheetname):
        """
        **** Marked for depreciation ****
        """
        #self.sheet = self.wb.get_sheet_by_name(sheetname)
        self.sheet = self.wb[sheetname]

    
    def get_column_count(self, sheetname):
        """
        Return the column count of the given sheet
        Example:
        | Get Column count     |  Sheet1 |
        """
        #self.sheet = self.wb.get_sheet_by_name(sheetname)
        self.sheet = self.wb[sheetname]
        return self.sheet.max_column


    def get_row_count(self, sheetname):
        """
        Return the Row count of the given sheet
        Example:
        | Get Row count     |  Sheet1 |
        """
        #self.sheet = self.wb.get_sheet_by_name(sheetname)
        self.sheet = self.wb[sheetname]
        return self.sheet.max_row

    def read_cell_data_by_coordinates(self,sheetname, row_value, column_value):
        """
        Return the value of a cell by giving the sheetname, row value & column value
        Example:
        | Read Cell Data By Coordinates     |  SheetName | Row Number |  Column Number  |
        | Read Cell Data By Coordinates     |  Sheet1 |  1  |  1  |
        """
        #self.sheet = self.wb.get_sheet_by_name(sheetname)
        self.sheet = self.wb[sheetname]
        self.row = int(row_value)
        self.column = int(column_value)
        varcellValue =  self.sheet.cell(row=self.row, column=self.column).value
        return varcellValue

    
    def write_data_by_coordinates(self,sheetname,row_value, column_value,varValue):
        """
        Write the value to a call using its co-ordinates
        Example:
        | Write Data By Coordinates    |  SheetName  | Row Number | Column Number |  Data  |
        | Write Data By Coordinates    | Sheet1 | 1 | 1 |  TestData  |
        """
        #self.sheet = self.wb.get_sheet_by_name(sheetname)
        self.sheet = self.wb[sheetname]
        self.row = int(row_value)
        self.column = int(column_value)
        self.varValue = varValue
        self.sheet.cell(row=self.row, column=self.column).value = self.varValue
    

    def save_excel(self, file):
        """
        Save the excel file after writing the data.
        Example:
        Update existing file:

        | Openexcel File       |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |
        | Save Excelfile       |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |

        Save in new file:
        | Openexcel File       |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |
        | Save Excelfile       |  D:\\Test\\ExcelRobotNewFile.xlsx                   |   
        """
        self.file = file
        self.wb.save(self.file)

    def add_new_sheet(self, varnewsheetname):
        """
        Add new sheet
        Arguments:
        | New sheetname        | The name of the new sheet to be added in the workbook     |

        Example:
        | Keywords             | Parameters                                       |
        | Add new sheet        | SheetName                                       |
        """
        self.newsheet = varnewsheetname
        self.wb.create_sheet(self.newsheet)